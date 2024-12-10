#
# Copyright (C) 2018 Pico Technology Ltd. See LICENSE file for terms.
#

import numpy as np
import openpyxl
from math import floor, log2, log10

def dataImporter(name):

    workbook = openpyxl.load_workbook(name, data_only = True)
    
    sheet = workbook['filterParameters']
    
    noOfChannels = sheet.cell(row = 2, column = 1).value
    bits = sheet.cell(row = 2, column = 2).value
    samplingRate = sheet.cell(row = 2, column = 3).value
    sampleLength = sheet.cell(row = 2, column = 4).value
    
    return noOfChannels, bits, samplingRate, sampleLength
    
    
def ps6000aTimebase(samplingRate):
    
    sampleInterval = (1/samplingRate)/1000000 #s
    
    breakPoint = 6.4/1000000000
    
    if sampleInterval >= breakPoint:
        timebase = floor((sampleInterval * 156250000)+5)
    else:
        timebase = floor(log2(sampleInterval * 5000000000))
    
    return timebase
    
def ps5000aTimebase(samplingRate):
    
    sampleInterval = (1/samplingRate)/1000000 #s
    
    breakPoint = 8/1000000000
    
    if sampleInterval >= breakPoint:
        timebase = floor((sampleInterval * 125000000)+2)
    else:
        timebase = floor(log2(sampleInterval * 1000000000))
    
    return timebase
    
def ps3000aTimebase(samplingRate):
    
    sampleInterval = (1/samplingRate)/1000000 #s
    
    breakPoint = 8/1000000000
    
    if sampleInterval >= breakPoint:
        timebase = floor((sampleInterval * 125000000)+2)
    else:
        timebase = floor(log2(sampleInterval * 1000000000))
    
    return timebase
  
def ps4000aTimebase(samplingRate):
    
    timebase = floor((80/samplingRate)-1)
    
    return timebase
    
def ps2000aTimebase(samplingRate):
    
    sampleInterval = (1/samplingRate)/1000000 #s
    
    breakPoint = 4/1000000000
    
    if sampleInterval>= breakPoint:
        timebase = floor((sampleInterval*125000000)+2)
    else:
        timebase = floor(log2(sampleInterval * 1000000000))
        
    return timebase
    
def ps2000Timebase(sampleRate):
    #assumes sample rate is in Hz
    #assumes sample interval in s
    
    sampleInterval = (1/sampleRate)
    
    timebase = floor(log10(sampleInterval*1000000000))
    
    return timebase
    
    
def BitEnumSelector(bits):

    if bits <= 8:
        enum = 0
    elif bits <= 10:
        enum = 10
    else:
        enum = 1
        
    return enum
    
def saveConfigFile(channels, bits, sampleRate,captureLength, maxAmplitude, segments):
    
    configValues = [channels, bits, sampleRate, captureLength, maxAmplitude, segments]
    
    # Save the list to a text file
    with open('configValues.txt', 'w') as file:
    # Write each element of the list on a new line
        for value in configValues:
            file.write(f"{value}\n")
        
    return
    
def loadConfigValues():
    
    restored_configValues = []
    
    with open('configValues.txt', 'r') as file:
        for line in file:
            value = line.strip()
            # Convert to integer or float as necessary
            if '.' in value:
                restored_configValues.append(float(value))
            else:
                restored_configValues.append(int(value))
            
    channels = restored_configValues[0]
    bits = restored_configValues[1]
    sampleRate = restored_configValues[2]
    captureLength = restored_configValues[3]
    maxAmplitude = restored_configValues[4]
    segments = restored_configValues[5]
    
    return channels, bits, sampleRate, captureLength, maxAmplitude, segments